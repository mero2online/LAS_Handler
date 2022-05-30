import os
import shutil
import lasio
import openpyxl
import xlwt
import xlrd
from copy2 import copy2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pandas import DataFrame
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from my_const import *
from HelperFunc import getFinalWellDate, resource_path, readLocalFile, writeLocalFile
from NewCurvesData import newPerCurves, newPerLithCurves, modPerCurves, newPerCurvesDSG
from GetFunc import convertNULL, GET_LITHO_EMPTY, Get_DSG_Formula, GetDSG_LAS_Header, getNewPerWellDSG, GetDSG_LAS_Header_ColorCode


def gen_litho_Percent_LAS(filename, start_depth):
    las = lasio.read(filename)

    wellNameOriginal = las.well.WELL.value
    finalWellName = wellNameOriginal[wellNameOriginal.find(
        '(')+len('('):wellNameOriginal.rfind(')')]
    las.well.WELL = finalWellName

    finalWellDate = getFinalWellDate()
    las.well.DATE = finalWellDate

    las.well.SRVC = 'EXLOG'

    # Remove Unused Curves
    for idx, x in enumerate(las.keys()):
        res = convertNULL(las[x])
        las.delete_curve(x)
        las.append_curve(x, res, descr=x)

    for idx, x in enumerate(las.keys()):
        if (idx == 21 or idx == 29):
            las.delete_curve(x)

    las.write(resource_path('draft_lithology_draft.las'),
              fmt='%.0f', len_numeric_field=5)

    for idx, x in enumerate(las.keys()):
        if (idx == 28 or idx == 29 or idx == 30 or idx == 31):
            las.delete_curve(x)

    # Convert NULL values and delete curve then append it
    for idx, x in enumerate(las.keys()):
        res = las[x]
        las.delete_curve(x)
        las.append_curve(newPerCurves[idx],
                         res, descr=newPerCurves[idx])

    firstRow = ' '.join(las.keys())
    lasFilename = resource_path('draft.las')
    excelFilename = resource_path('draft.xlsx')
    csvFilename = resource_path('draft.txt')

    las.write(lasFilename, fmt='%.0f', len_numeric_field=5)
    las.to_excel(excelFilename)
    las.to_csv(csvFilename, units=False, delimiter='\t')
    csvDraft = readLocalFile(csvFilename)
    csvDraftWithoutDecimal = csvDraft.replace('.0', '')
    writeLocalFile(resource_path(
        f'out\\DSG_FOR_GRAVITAS_CONVERTER.txt'), csvDraftWithoutDecimal)

    DSG()
    LITHOLOGY()

    trimLASandEXCEL(lasFilename, excelFilename, firstRow)

    lithology_gravitas_name = f'{finalWellName}_LITHOLOGY_{finalWellDate}_GRAVITAS'
    shutil.copy(resource_path('draft.las'), resource_path(
        f'out\\{lithology_gravitas_name}.las'))

    lasDraft = readLocalFile(lasFilename)
    lasDraftSplitted = lasDraft.splitlines()

    all_rows = []
    for idx, x in enumerate(lasDraftSplitted):
        if idx > 0:
            x = [int(i) for i in x.split()]
            all_rows.append(x)

    # load the excel file
    inBook = xlrd.open_workbook(
        resource_path('LITHOLOGY_GRAVITAS.xls'), formatting_info=True, on_demand=True)
    inSheet = inBook .sheet_by_index(0)

    # copy the contents of excel file
    outBook, outStyle = copy2(inBook)

    # open the first sheet
    w_sheet = outBook.get_sheet(0)

    for idx, row in enumerate(all_rows, 1):
        for idc, cell in enumerate(row):
            xf_index = inSheet.cell_xf_index(idx, idc)
            saved_style = outStyle[xf_index]
            w_sheet.write(idx, idc, cell, saved_style)

    outBook.get_sheet(0).name = f'{finalWellName}_LITHOLOGY_GRAVITAS'
    # save the file
    outBook.save(resource_path(
        f'out\\{lithology_gravitas_name}.xls'))

    if (start_depth):
        LITHOLOGY_GRAVITAS_Converted(
            lasFilename, finalWellName, finalWellDate, start_depth)

#
# LITHOLOGY_GRAVITAS_Converted
#


def LITHOLOGY_GRAVITAS_Converted(lasFilename, finalWellName, finalWellDate, start_depth):
    lasDraft = readLocalFile(lasFilename)
    lasDraftSplitted = lasDraft.splitlines()

    firstNumRow = [int(i) for i in lasDraftSplitted[1].split()]
    filtered = []
    filtered.append(firstNumRow)
    for idx, x in enumerate(lasDraftSplitted):
        lastFiltered = filtered[len(filtered)-1]
        if idx > 0:
            x = [int(i) for i in x.split()]
            if x[1:len(x)-1] != lastFiltered[1:len(lastFiltered)-1]:
                filtered.append(x)
            else:
                filtered.pop()
                filtered.append(x)

    filtered.insert(0, [lasDraftSplitted[0].split()])

    result = []
    for idx, row in enumerate(filtered):
        if (idx == 0):
            headerRow = row[0]
            # headerRow[1:len(headerRow)] = headerRow[len(headerRow)-1:0:-1]
        else:
            matched = []
            # row[1:len(row)] = row[len(row)-1:0:-1]
            for id, val in enumerate(row):
                if id > 0:
                    if val > 0:
                        matched.append(val)
                        top = getTop(result, row, idx, int(start_depth))
                        base = row[0]
                        right = val if len(matched) <= 1 else sum(matched)
                        left = right - val
                        result.append([top, base, left, right, headerRow[id]])

    result.insert(0, ['TOP', 'BASE', 'LEFT', 'RIGHT', 'LITHOTYPE'])

    finalFileName = f'{finalWellName}_LITHOLOGY_GRAVITAS_{finalWellDate}_Converted'
    wb = Workbook()
    ws1 = wb.active

    for row in result:
        ws1.append(row)
    wb.save(resource_path(f'out\\{finalFileName}.xlsx'))

    df = DataFrame(ws1.values)
    df.to_csv(resource_path(f'out\\{finalFileName}.txt'),
              index=False, header=False, sep='\t')
    #
    # DSG
    #


def DSG():
    las = lasio.read(resource_path('draft.las'))

    # Convert all 'SNDSH' values to 0 and Change position of 'CEMENT'
    for idx, x in enumerate(las.keys()):
        if (idx == 18):
            data = [0]*len(las[x])
            las.delete_curve(x)
            las.insert_curve(18, x, data)
        elif (idx == 27):
            data = las[x]
            las.delete_curve(x)
            las.insert_curve(26, x, data)

    # Remove Unused Curves
    for idx, x in enumerate(las.keys()):
        if (idx == 9 or idx == 10 or idx == 16 or idx == 25):
            las.delete_curve(x)

    lasFilename = resource_path('draft_DSG.las')
    excelFilename = resource_path('draft_DSG.xlsx')

    las.to_excel(excelFilename)

    # Rename all curves for DSG by delete curve then insert it
    for idx, x in enumerate(las.keys()):
        mnemonic = newPerCurvesDSG[idx]['name']
        data = las[x] if idx <= 1 else las[x]+las[las.keys()[idx-1]]
        unit = 'F' if idx < 1 else '%'
        descr = newPerCurvesDSG[idx]['desc']
        las.delete_curve(x)
        las.insert_curve(idx, mnemonic, data, unit=unit, descr=descr)

    workbook = openpyxl.load_workbook(excelFilename)
    wsh = workbook['Curves']

    cellAlignment = Alignment(horizontal='center', vertical='center')
    for idx, cell in enumerate(wsh["1:1"]):
        col_letter = get_column_letter(idx+1)
        fg = GetDSG_LAS_Header_ColorCode(idx)
        cell.value = las.keys()[idx]
        wsh.column_dimensions[col_letter].width = len(f"{cell.value}") * 1.5
        cell.alignment = cellAlignment
        cell.font = Font(name='Courier New')
        if fg:
            cell.fill = PatternFill(fgColor=fg,  fill_type="solid")
        for idx, cell in enumerate(wsh[col_letter]):
            cell.alignment = cellAlignment

    workbook.save(resource_path(excelFilename))

    las.insert_curve(
        1, 'DEPTH_ORIG', las['DEPTH'], unit='F', descr='Depth Orig')

    las.well['STEP'].descr = 'STEP DEPTH'
    del las.well['PROV']

    cwd = os.getcwd()
    text = readLocalFile(f'{cwd}\LAS_Handler_DSG_Config.csv')

    result = []

    for line in text.splitlines():
        result.append(line.split(",")[1])

    reAdd = ['UWI', 'WELL', 'COMP', 'FLD', 'LOC', 'SRVC', 'DATE']
    for x in reAdd:
        result.append(las.well[x].value)
        del las.well[x]

    newPerWellDSG = getNewPerWellDSG(result)
    for idx, x in enumerate(newPerWellDSG):
        las.well[newPerWellDSG[idx]['mnemonic']] = lasio.HeaderItem(
            mnemonic=newPerWellDSG[idx]['mnemonic'],
            value=newPerWellDSG[idx]['value'],
            descr=newPerWellDSG[idx]['descr'],
            unit=newPerWellDSG[idx]['unit'],)

    las.write(lasFilename, fmt='%.0f', len_numeric_field=5)

    txt = readLocalFile(lasFilename)
    startOne = '~Well ------------------------------------------------------'
    endOne = '~Curve Information -----------------------------------------'
    textOneLas = txt[txt.find(startOne)+len(startOne):txt.rfind(endOne)]

    startTwo = '~Curve Information -----------------------------------------'
    endTwo = '~Params ----------------------------------------------------'
    textTwoLas = txt[txt.find(startTwo)+len(startTwo):txt.rfind(endTwo)]

    finalDSG_Header = GetDSG_LAS_Header(textOneLas, textTwoLas)

    firstRow = '~A '+' '.join(las.keys())
    trimLASandEXCEL(lasFilename, excelFilename, firstRow)

    workbook = openpyxl.load_workbook(excelFilename)
    ws2Title = f"{las.well['WELL'].value}_LITHOLOGY-DSG"
    workbook.create_sheet(title=ws2Title)
    ws = workbook['Curves']
    ws.title = 'original values'
    ws1 = workbook['original values']
    ws2 = workbook[ws2Title]

    ws1.freeze_panes = ws1['A2']

    df = DataFrame(ws1.values)
    rows = dataframe_to_rows(df, index=False, header=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            values = Get_DSG_Formula(r_idx)
            ws2.cell(row=r_idx, column=c_idx,
                     value=value if r_idx == 1 else values[c_idx-1])

    ws2.insert_cols(2, 1)
    for x in range(len(ws1['A'])):
        i = x+1
        if x == 0:
            ws2[f'B{i}'] = 'Depth.org'
        else:
            ws2[f'B{i}'] = f'=A{i}'

    for idx, cell in enumerate(ws2["1:1"]):
        col_letter = get_column_letter(idx+1)
        ws2.column_dimensions[col_letter].width = len(f"{cell.value}") * 1.5
        cell.alignment = cellAlignment
        cell.font = Font(name='Courier New')
        for idx, cell in enumerate(ws2[col_letter]):
            cell.alignment = cellAlignment

    finalFileNameXlsx = f'{las.well.WELL.value}_LITHOLOGY-DSG_GRAVITAS_{las.well.DATE.value}'
    finalFileNameLas = f'{las.well.WELL.value}_LITHOLOGY-DSG_{las.well.DATE.value}'
    # workbook.save(excelFilename)
    workbook.save(resource_path(f'out\\{finalFileNameXlsx}.xlsx'))

    finalData = readLocalFile(lasFilename)
    finalLAS = f'{finalDSG_Header}{finalData}'
    writeLocalFile(resource_path(f'out\\{finalFileNameLas}.las'), finalLAS)

#
# LITHOLOGY
#


def LITHOLOGY():
    las = lasio.read(resource_path('draft.las'))
    lasLithology = lasio.read(resource_path('draft_lithology_draft.las'))
    newLas = lasio.LASFile()
    newLas.add_curve('DEPTH', las['DEPTH'], unit='ft', descr='1 Hole Depth')

    for idx, x in enumerate(newPerLithCurves):
        if idx <= 19:
            newLas.add_curve(newPerLithCurves[idx]['name'], las[modPerCurves[idx]],
                             unit='%', descr=newPerLithCurves[idx]['desc'])
        else:
            newLas.add_curve(newPerLithCurves[idx]['name'],
                             lasLithology[lasLithology.keys()[idx+8]], unit='%', descr=newPerLithCurves[idx]['desc'])

    lasFilename = resource_path('draft_LITHOLOGY.las')
    excelFilename = resource_path('draft_LITHOLOGY.xlsx')
    firstRow = ' '.join(newLas.keys())

    newLas.write(lasFilename, fmt='%.0f', len_numeric_field=5)
    newLas.to_excel(excelFilename)

    txt = readLocalFile(resource_path('draft.las'))
    startOne = '~Well ------------------------------------------------------'
    endOne = '~Curve Information -----------------------------------------'
    textOneLas = txt[txt.find(startOne)+len(startOne):txt.rfind(endOne)]
    lith = readLocalFile(resource_path('draft_LITHOLOGY.las'))
    startTwo = '~Curve Information -----------------------------------------'
    endTwo = '~Params ----------------------------------------------------'
    textTwoLas = lith[lith.find(startTwo)+len(startTwo):lith.rfind(endTwo)]
    startThree = '~ASCII -----------------------------------------------------'
    textThreeLas = lith[lith.find(startThree)+len(startThree):len(lith)-1]
    finalData = f'{text1}{textOneLas}{text4}{textTwoLas}{text5}{textThreeLas}'

    finalFileName = f'{las.well.WELL.value}_LITHOLOGY_{las.well.DATE.value}'
    writeLocalFile(resource_path(f'out\\{finalFileName}.las'), finalData)

    wb = Workbook()
    ws1 = wb.active
    data = finalData.splitlines()
    for idx, row in enumerate(data):
        if idx <= 55:
            ws1.append([row])
        elif idx == 56:
            one = row.split()
            two = one[2:len(one)]
            two.insert(0, ' '.join([one[0], one[1]]))
            ws1.append(two)
        else:
            ws1.append([int(x) for x in row.split()])

    wb.save(resource_path(f'out\\{finalFileName}.xlsx'))

    trimLASandEXCEL(lasFilename, excelFilename, firstRow)


# ########
# Helper #
# ########


def trimLASandEXCEL(lasFilename, excelFilename, firstRow):
    workbook = openpyxl.load_workbook(excelFilename)
    std = workbook.get_sheet_by_name('Header')
    workbook.remove_sheet(std)
    workbook.save(excelFilename)

    txt = readLocalFile(lasFilename)

    start = '~ASCII -----------------------------------------------------'
    data = txt[txt.find(start)+len(start):len(txt)-1]
    finalData = f'{firstRow}{data}'

    writeLocalFile(lasFilename, finalData)


def getTop(result, row, idx, start_depth):
    if idx == 1:
        return start_depth
    if (len(result) > 0 and ((row[0] - result[len(result)-1][1]) > 10)):
        return result[len(result)-1][1]
    elif (len(result) > 0 and result[len(result)-1][1]-result[len(result)-1][0] > 10) and row[0] == result[len(result)-1][1]:
        return result[len(result)-1][0]
    else:
        return row[0] - 10
