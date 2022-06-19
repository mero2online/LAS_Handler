import lasio
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

from my_const import *
from GetFunc import aggregate_ROP_FiveFeet, aggregate_DEPTH_FiveFeet
from HelperFunc import getFinalWellDate, resource_path, readLocalFile, writeLocalFile


def gen_ROP_LAS(filename):
    las = lasio.read(filename)

    wellNameOriginal = las.well.WELL.value
    finalWellName = wellNameOriginal[wellNameOriginal.find(
        '(')+len('('):wellNameOriginal.rfind(')')]
    las.well.WELL = finalWellName

    finalWellDate = getFinalWellDate()
    las.well.DATE = finalWellDate

    las.well.SRVC = 'EXLOG'

    las.append_curve('DEPTH', aggregate_DEPTH_FiveFeet(las['DMEA']),
                     unit='FT', descr='Depth')

    las.append_curve('ROP5_ML', aggregate_ROP_FiveFeet(las['ROPA']),
                     unit='MIN/5FT', descr='Rate Of Penetration')

    las.delete_curve('DMEA')
    las.delete_curve('ROPA')

    las.write(resource_path('draft.las'), fmt='%.0f', len_numeric_field=5)

    txt = readLocalFile(resource_path('draft.las'))

    startOne = '~Well ------------------------------------------------------'
    endOne = '~Curve Information -----------------------------------------'
    textOneLas = txt[txt.find(startOne)+len(startOne):txt.rfind(endOne)]

    startTwo = '~Curve Information -----------------------------------------'
    endTwo = '~Params ----------------------------------------------------'
    textTwoLas = txt[txt.find(startTwo)+len(startTwo):txt.rfind(endTwo)]

    startThree = '~ASCII -----------------------------------------------------'
    textThreeLas = txt[txt.find(startThree)+len(startThree):len(txt)-1]

    finalData = f'{text1}{textOneLas}{text6}{textTwoLas}{text7}{textThreeLas}'

    writeLocalFile(resource_path('draft.las'), finalData)

    finalFileNameLas = f'{las.well.WELL.value}_ROP-DSG_{las.well.DATE.value}'
    finalFileNameXlsx = f'{las.well.WELL.value}_ROP_{las.well.DATE.value}_GRAVITAS'
    writeLocalFile(resource_path(f'out\\{finalFileNameLas}.las'), finalData)

    wb = Workbook()
    ws1 = wb.active
    data = finalData.splitlines()
    ws1.append(['Depth (ft)', 'ROP (min/5 ft)'])

    for idx, row in enumerate(data):
        if idx > 33:
            ws1.append([int(x) for x in row.split()])

    columnToStyle = ['A', 'B']

    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 18

    ws1.row_dimensions[1].height = 27

    topCellFont = Font(size=14, bold=True)
    cellFont = Font(bold=True)

    borderStyle = Side(border_style='thin', color='000000')
    borderPosition = Border(
        left=borderStyle, right=borderStyle, top=borderStyle, bottom=borderStyle)

    cellAlignment = Alignment(horizontal='center', vertical='center')

    for col in columnToStyle:
        for idx, cell in enumerate(ws1[col]):
            cell.font = topCellFont if idx == 0 else cellFont
            cell.border = borderPosition
            cell.alignment = cellAlignment

    ws1.freeze_panes = ws1['A2']

    wb.save(resource_path(f'out\\{finalFileNameXlsx}.xlsx'))

    depth = list(las['DEPTH'])
    rop = list(las['ROP5_ML'])

    rangeToAdd = int(depth[0] % 1000/5-1)
    start = int("{:.0f}".format(depth[0]/1000))*1000
    for i in range(rangeToAdd):
        valueToInsert = start + (i+1)*5
        depth.insert(i, float(valueToInsert))
        rop.insert(i, '')

    depthChunks = [depth[i:i + 20] for i in range(0, len(depth), 20)]
    ropChunks = [rop[i:i + 20] for i in range(0, len(rop), 20)]
    ropChunksSheets = [ropChunks[i:i + 10]
                       for i in range(0, len(ropChunks), 10)]

    rightCol = []
    rightCol.append('')
    for idx in range(1, 21, 1):
        rightCol.append(f'{idx*5-5}-{idx*5}')
    for idx, v in enumerate(ropChunksSheets):
        ropChunksSheets[idx].insert(0, rightCol)

    wbDT = Workbook()

    for i, v in enumerate(ropChunksSheets):
        toDepth = start + (i+1)*1000
        fromDepth = toDepth-1000
        for idx, m in enumerate(v):
            if idx > 0:
                m.insert(0, fromDepth+((idx-1)*100))
        ws2Title = f'{fromDepth}-{toDepth}'
        wbDT.create_sheet(title=ws2Title)
        ws2 = wbDT[ws2Title]
        for c_idx, col in enumerate(v, 1):
            for r_idx, value in enumerate(col, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)

    std = wbDT.get_sheet_by_name('Sheet')
    wbDT.remove_sheet(std)
    wbDT.save(resource_path(f'out\\{las.well.WELL.value}_DRILL_TIME.xlsx'))
